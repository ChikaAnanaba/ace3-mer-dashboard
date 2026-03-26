"""
ACE3 Excel Export
Generates ACE3_Data_Report.xlsx with auto-populated DATA_INPUT sheet.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from targets import INDICATOR_META, qoq_change, status_flag

# ── style helpers ─────────────────────────────────────────────────────────────
NAVY  = '1F4E79'; BLUE  = '2E75B6'; GREEN = '1A6632'; RED   = 'BA0C2F'
AMBER = 'C87000'; WHITE = 'FFFFFF'; LIGHT = 'F4F6F9'; LGREEN= 'E6F2EB'
LRED  = 'FDEAEA'; LAMBER= 'FEF3E2'; LBLUE = 'EBF3FB'; CALC  = 'F8FAFC'

def _fill(hex_): return PatternFill('solid', fgColor=hex_)
def _font(sz=10, bold=False, color='1E293B', name='Arial'):
    return Font(name=name, size=sz, bold=bold, color=color)
def _border():
    s = Side(style='thin', color='CBD5E1')
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_row(ws, row, values, bg=NAVY, fg=WHITE, height=22):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font  = Font(name='Arial', size=9, bold=True, color=fg)
        c.fill  = _fill(bg)
        c.alignment = _align('center')
        c.border = _border()
    ws.row_dimensions[row].height = height

def _write(ws, row, col, val, bg=WHITE, bold=False, color='1E293B',
           fmt=None, align='left', sz=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = _font(sz, bold, color)
    c.fill      = _fill(bg)
    c.alignment = _align(align)
    c.border    = _border()
    if fmt:
        c.number_format = fmt
    ws.row_dimensions[row].height = 17
    return c


# ── indicator rows definition ─────────────────────────────────────────────────
# (indicator_key, display_label, category, pepfar_code, MER_freq, has_target)
ROWS = [
    # Treatment
    ('TX_CURR',        'Active on ART (Snapshot)',            'Treatment',      'TX_CURR',       'Quarterly',   True),
    ('TX_NEW',         'New ART Initiations',                 'Treatment',      'TX_NEW',        'Quarterly',   True),
    ('TX_ML',          'Treatment Interruptions (IIT)',       'Retention',      'TX_ML',         'Quarterly',   False),
    ('TX_RTT',         'Returned to Treatment',               'Retention',      'TX_RTT',        'Quarterly',   False),
    # VL
    ('TX_PVLS_D',      'VL Tests Resulted (Denominator)',     'Viral Load',     'TX_PVLS_D',     'Quarterly',   True),
    ('TX_PVLS_N',      'VL Suppressed (Numerator)',           'Viral Load',     'TX_PVLS_N',     'Quarterly',   True),
    ('VL_COVERAGE',    'VL Coverage % (derived)',             'Viral Load',     'VL_COV',        'Quarterly',   False),
    ('VL_SUPPRESSION', 'VL Suppression % (derived)',          'Viral Load',     'VL_SUPP',       'Quarterly',   False),
    # Testing
    ('HTS_TST',        'HIV Tests Conducted',                 'Testing',        'HTS_TST',       'Quarterly',   True),
    ('HTS_TST_POS',    'HIV Positive Results',                'Testing',        'HTS_TST_POS',   'Quarterly',   True),
    # PMTCT
    ('PMTCT_STAT_N',   'ANC Clients Tested for HIV',         'PMTCT',          'PMTCT_STAT',    'Quarterly',   True),
    ('PMTCT_STAT_POS', 'ANC HIV Positive',                   'PMTCT',          'PMTCT_STAT_POS','Quarterly',   True),
    ('PMTCT_ART_D',    'PMTCT on ART',                       'PMTCT',          'PMTCT_ART_D',   'Quarterly',   True),
    ('PMTCT_EID',      'EID PCR Tests Done',                 'PMTCT',          'PMTCT_EID',     'Quarterly',   False),
    # TB
    ('TB_SCREEN',      'TB Screening (TX_CURR)',              'TB/HIV',         'TB_SCRN',       'Quarterly',   False),
    ('TB_PREV_N',      'TB Preventive Therapy (TPT)',         'TB/HIV',         'TB_PREV_N',     'Semi-Annual', True),
    ('TX_TB_N',        'TB/HIV Co-infected on ART',           'TB/HIV',         'TX_TB_N',       'Semi-Annual', True),
    ('TB_ART',         'TB Patients on ART',                  'TB/HIV',         'TB_ART',        'Annual',      True),
    # PrEP
    ('PrEP_NEW',       'New PrEP Initiations',               'PrEP',           'PrEP_NEW',      'Quarterly',   True),
    ('PrEP_CT',        'PrEP Continuing (Follow-up visits)', 'PrEP',           'PrEP_CT',       'Quarterly',   False),
    ('PrEP_CURR',      'Currently on PrEP (Snapshot)',       'PrEP',           'PrEP_CURR',     'Quarterly',   False),
    # DSD
    ('MMD_3P',         'MMD ≥3 Months',                     'DSD',            'MMD_3P',        'Quarterly',   False),
    ('MMD_6P',         'MMD 6+ Months',                      'DSD',            'MMD_6P',        'Quarterly',   False),
    # CxCa
    ('CXCA_SCRN',      'CxCa Eligible WLHIV',               "Women's Health", 'CXCA_SCRN',     'Semi-Annual', False),
    ('CXCA_TX',        'CxCa Screened',                      "Women's Health", 'CXCA_TX',       'Semi-Annual', False),
    # AHD
    ('AHD_SCRN',       'AHD Screened',                       'AHD',            'AHD_SCRN',      'Quarterly',   False),
    ('AHD_CONF',       'AHD Confirmed',                      'AHD',            'AHD_CONF',      'Quarterly',   False),
]

SNAPSHOT_KEYS = {'TX_CURR', 'VL_COVERAGE', 'VL_SUPPRESSION', 'PrEP_CURR', 'MMD_3P', 'MMD_6P'}
PCT_KEYS      = {'VL_COVERAGE', 'VL_SUPPRESSION', 'HTS_YIELD'}


def generate_excel(results: dict) -> bytes:
    """
    Build ACE3_Data_Report.xlsx and return as bytes.

    Parameters
    ----------
    results : dict from assembler.assemble()
    """
    q1   = results.get('q1', {})
    q2   = results.get('q2', {})
    semi = results.get('semi', {})
    tgts = results.get('targets', {})

    wb = Workbook()

    # ── SHEET 1: DATA INPUT ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'DATA_INPUT'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A4'

    # Title block
    ws.merge_cells('A1:R1')
    t = ws['A1']
    t.value = 'ACE3 PROGRAMME  ·  INDICATOR DATA REPORT  —  AUTO-GENERATED'
    t.font  = Font(name='Arial', size=13, bold=True, color=WHITE)
    t.fill  = _fill(NAVY)
    t.alignment = _align('center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:R2')
    s = ws['A2']
    s.value = 'All values auto-populated from uploaded line lists. Shaded cells are derived / calculated.'
    s.font  = Font(name='Arial', size=9, italic=True, color='475569')
    s.fill  = _fill('F1F5F9')
    s.alignment = _align('left')
    ws.row_dimensions[2].height = 15

    # Column headers
    headers = [
        'Indicator', 'Category', 'PEPFAR Code', 'MER Frequency',
        'Q1 Actual', 'Q2 Actual', 'Semi-Annual Total',
        'Q1→Q2 Change', 'Q1→Q2 %Δ',
        'Q3 Actual', 'Q4 Actual', 'Q3→Q4 Change', 'Q3→Q4 %Δ',
        'Annual Target', 'Q1 % of Target', 'Q2 % of Target (cum)',
        'Semi-Annual %', 'Status'
    ]
    _hdr_row(ws, 3, headers, bg=NAVY, height=30)

    # Column widths
    col_widths = [38, 16, 14, 13, 11, 11, 15, 13, 10, 11, 11, 13, 10, 13, 14, 16, 13, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for idx, (key, label, cat, code, freq, has_tgt) in enumerate(ROWS):
        row = 4 + idx
        row_bg = WHITE if idx % 2 == 0 else 'F8FAFC'

        v1   = q1.get(key)
        v2   = q2.get(key)
        vsemi = semi.get(key)
        tgt  = tgts.get(INDICATOR_META.get(key, {}).get('target_col', ''))

        # Q1→Q2 change
        abs_chg, pct_chg = qoq_change(
            v1 if isinstance(v1, (int, float)) else None,
            v2 if isinstance(v2, (int, float)) else None
        )

        # % of target
        def _pct_tgt(val):
            if val is None or tgt is None or tgt == 0: return None
            return round(val / tgt * 100, 1)

        q1_pct  = _pct_tgt(v1)
        q2_cum  = None
        if v1 is not None and v2 is not None and not key in SNAPSHOT_KEYS:
            q2_cum = _pct_tgt((v1 if isinstance(v1,(int,float)) else 0) +
                               (v2 if isinstance(v2,(int,float)) else 0))
        elif key in SNAPSHOT_KEYS:
            q2_cum = _pct_tgt(v2)
        semi_pct = _pct_tgt(vsemi)

        # status from semi-annual, fallback q2
        s_pct = semi_pct if semi_pct is not None else q2_cum
        status_txt = status_flag(s_pct) if has_tgt else '—'

        # Format values
        is_pct = key in PCT_KEYS
        num_fmt = '0.0%' if is_pct else '#,##0'
        chg_fmt = '+#,##0;-#,##0;0'

        # Write cells
        _write(ws, row, 1,  label,  row_bg, True, '0F172A', sz=9)
        _write(ws, row, 2,  cat,    row_bg, False, '475569', sz=9)
        _write(ws, row, 3,  code,   row_bg, True,  NAVY, sz=9)
        _write(ws, row, 4,  freq,   row_bg, False, '64748B', sz=9)

        # Q1
        c1 = _write(ws, row, 5, v1 if isinstance(v1,(int,float)) else None,
                    LBLUE if v1 is not None else row_bg, True, '000080', fmt=num_fmt, align='right')

        # Q2
        c2 = _write(ws, row, 6, v2 if isinstance(v2,(int,float)) else None,
                    LBLUE if v2 is not None else row_bg, True, '000080', fmt=num_fmt, align='right')

        # Semi-annual total
        sv = vsemi if isinstance(vsemi,(int,float)) else None
        _write(ws, row, 7, sv, CALC, False, '000000', fmt=num_fmt, align='right')

        # Q1→Q2 change
        if abs_chg is not None:
            chg_color = RED if abs_chg < 0 else GREEN
            _write(ws, row, 8, abs_chg, row_bg, True, chg_color, fmt=chg_fmt, align='right')
            _write(ws, row, 9, pct_chg/100 if pct_chg is not None else None,
                   row_bg, True, chg_color, fmt='+0.0%;-0.0%;0%', align='right')
        else:
            for col in [8, 9]:
                _write(ws, row, col, '—', row_bg, False, '94A3B8', align='center')

        # Q3 / Q4 (placeholders)
        for col in [10, 11]:
            _write(ws, row, col, None, row_bg, False, '94A3B8', align='center')
        for col in [12, 13]:
            _write(ws, row, col, '—', row_bg, False, '94A3B8', align='center')

        # Annual target
        if has_tgt and tgt:
            _write(ws, row, 14, tgt, LAMBER, True, '7B4500', fmt='#,##0', align='right')
        else:
            _write(ws, row, 14, 'No target', row_bg, False, '94A3B8', align='center')

        # % of target columns
        pct_fmt = '0.0%'
        for col, pct_val in [(15, q1_pct), (16, q2_cum), (17, semi_pct)]:
            if pct_val is not None and has_tgt:
                bg = LGREEN if pct_val >= 75 else (LAMBER if pct_val >= 50 else LRED)
                fc = GREEN  if pct_val >= 75 else (AMBER  if pct_val >= 50 else RED)
                _write(ws, row, col, pct_val/100, bg, True, fc, fmt=pct_fmt, align='right')
            else:
                _write(ws, row, col, '—', row_bg, False, '94A3B8', align='center')

        # Status
        if has_tgt and s_pct is not None:
            s_bg = LGREEN if '✓' in status_txt else (LAMBER if '⚠' in status_txt else LRED)
            s_fc = GREEN  if '✓' in status_txt else (AMBER  if '⚠' in status_txt else RED)
            _write(ws, row, 18, status_txt, s_bg, True, s_fc, fmt=None, align='center')
        else:
            _write(ws, row, 18, '—', row_bg, False, '94A3B8', align='center')

    # Legend row
    leg_row = 4 + len(ROWS) + 1
    ws.merge_cells(f'A{leg_row}:R{leg_row}')
    leg = ws[f'A{leg_row}']
    leg.value = ('KEY:  Blue cells = auto-filled actuals   |   Amber = annual target   |   '
                 '✓ On Track ≥75%   |   ⚠ Watch 50–74%   |   ✗ Behind <50%   |   '
                 'Snapshot indicators (TX_CURR, PrEP_CURR) show period-end value, not cumulative')
    leg.font      = Font(name='Arial', size=8, italic=True, color='475569')
    leg.fill      = _fill(LBLUE)
    leg.alignment = _align('left')
    ws.row_dimensions[leg_row].height = 15

    # ── save to bytes ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
